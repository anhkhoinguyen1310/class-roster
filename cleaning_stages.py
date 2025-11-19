"""
School-specific data cleaning stages.

Extend this module with custom DataCleaningStage implementations for each school format.
Each school sends data in a different format, so we normalize it to a standard format:
- Columns: Teacher, Student Name, Class (optionally Advisor)
- Clean data: trimmed, no duplicates, proper case
"""

from class_roster_ui import DataCleaningStage, ProcessingPipeline, ClassSplittingStage
from openpyxl import Workbook


class ROCLCleaningStage(DataCleaningStage):
    """
    Universal cleaning for ROCL-style rosters.

    Goals:
    - Auto-detect columns instead of assuming fixed positions.
    - Support:
      * Name + Class + Teacher
      * Class + First Name + Last Name (no teacher)
      * Multiple sheets (combines all into one)
    - Output: Teacher | Student Name | Class (plus Advisor if requested).
    """

    # Header candidates for each logical field
    CLASS_HEADERS = [
        "class", "class id", "class#", "ocl",
        "class section id", "class section", "section",
        "grade", "grade level"
    ]
    STUDENT_FULL_HEADERS = [
        "name", "student", "student name",
        "full name", "student full", "stu name"
    ]
    STUDENT_FIRST_HEADERS = [
        "first name", "fname", "student first name", "given name", "first", "fname"
    ]
    STUDENT_LAST_HEADERS = [
        "last name", "lname", "student last name", "surname", "last", "lname"
    ]
    TEACHER_HEADERS = [
        "teacher", "teacher name", "homeroom teacher", "tchr"
    ]
    ADVISOR_HEADERS = [
        "advisor", "adv", "adviser"
    ]

    def __init__(self, include_advisor: bool = False):
        self.include_advisor = include_advisor

    def process(self, data: dict) -> dict:
        wb_in = data.get("workbook")

        if not wb_in:
            raise ValueError("Workbook not found")

        # First, extract teacher schedule mapping from Schedule sheet if it exists
        teacher_map = self._extract_teacher_map(wb_in)
        if teacher_map:
            print(f"[ROCL] Extracted teacher map with {len(teacher_map)} entries")

        # Collect records from ALL sheets
        all_cleaned_records = []
        
        print(f"[ROCL] Processing {len(wb_in.sheetnames)} sheets: {wb_in.sheetnames}")
        
        for sheet_name in wb_in.sheetnames:
            # Skip non-data sheets (e.g., Schedule, Instructions, etc.)
            if sheet_name.lower().strip() in ["schedule", "schedule ", "instructions", "notes"]:
                print(f"[ROCL] Skipping non-data sheet: {sheet_name}")
                continue
                
            print(f"[ROCL] Scanning sheet: {sheet_name}")
            ws = wb_in[sheet_name]
            sheet_records = self._extract_and_clean(ws)
            
            # Fill in teacher data from schedule map if available
            if teacher_map:
                for record in sheet_records:
                    class_val = record.get("class", "").strip()
                    teachers = self._find_teachers_for_class(class_val, teacher_map)
                    if teachers:
                        record["teacher"] = teachers
            
            all_cleaned_records.extend(sheet_records)
            print(f"[ROCL] Found {len(sheet_records)} records in {sheet_name}")

        # Create standardized output workbook with all records
        wb_out = self._create_cleaned_workbook(all_cleaned_records)

        data["workbook"] = wb_out
        data["worksheet"] = wb_out.active
        data["cleaned_record_count"] = len(all_cleaned_records)
        return data
    
    def _find_teachers_for_class(self, class_str, teacher_map):
        """
        Find teachers for a given class value.
        Handles both exact matches and grade-level matches.
        """
        # First try exact match
        if class_str in teacher_map:
            return teacher_map[class_str]
        
        # Extract grade number from class string (e.g., "6th Grade" -> "6")
        import re
        grade_match = re.search(r'(\d+)', class_str)
        if not grade_match:
            return ""
        
        grade_num = grade_match.group(1)
        
        # Find all teachers for this grade
        grade_teachers = []
        for class_key, teacher in teacher_map.items():
            if class_key.startswith(grade_num + ".") or f"Grade {grade_num}" in class_key:
                grade_teachers.append(teacher)
        
        # Return combined list of all teachers for this grade
        if grade_teachers:
            return " / ".join(grade_teachers)
        
        return ""
    
    def _extract_teacher_map(self, wb_in):
        """
        Extract class-to-teacher mapping from Schedule sheet.
        Returns dict: {class_id: teacher_name}
        """
        schedule_sheet = None
        for sheet_name in wb_in.sheetnames:
            if sheet_name.lower().strip() in ["schedule", "schedule "]:
                schedule_sheet = wb_in[sheet_name]
                break
        
        if not schedule_sheet:
            return {}
        
        teacher_map = {}
        # Schedule table: Class is in column B (index 1), Teacher is in column E (index 4)
        # Table starts at row 5, with header at row 4
        for row_idx in range(5, 50):
            row = list(schedule_sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
            
            class_val = row[1] if len(row) > 1 else None
            teacher_val = row[4] if len(row) > 4 else None
            
            if not class_val or not teacher_val:
                continue
            
            class_str = str(class_val).strip()
            teacher_str = str(teacher_val).strip()
            
            # Skip header rows and empty rows
            if not class_str or not teacher_str or class_str.lower() == "class":
                continue
            
            # Clean up teacher name (remove newlines, extra spaces)
            teacher_str = teacher_str.replace('\n', ' / ').replace('  ', ' ').strip()
            
            teacher_map[class_str] = teacher_str
            print(f"[ROCL] Mapped class '{class_str}' -> teacher '{teacher_str}'")
        
        return teacher_map

    def _find_header_row_and_indices(self, worksheet):
        """
        Find the header row and return column indices for key fields.
        Uses two-pass matching: exact match first, then substring match.
        Prioritizes rows with better header quality.
        
        Returns: (header_row_idx, {field: col_idx})
        """
        candidates = []
        
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=10)):
            headers = [cell.value for cell in row]
            headers_lower = [str(h).lower().strip() if h else "" for h in headers]
            
            # Two-pass matching: exact first, then substring
            def find_header(candidates_list):
                # Pass 1: exact match
                for candidate in candidates_list:
                    for idx, h in enumerate(headers_lower):
                        if h == candidate.lower():
                            return idx
                # Pass 2: substring match
                for candidate in candidates_list:
                    for idx, h in enumerate(headers_lower):
                        if candidate.lower() in h:
                            return idx
                return None
            
            student_full_idx = find_header(self.STUDENT_FULL_HEADERS)
            student_first_idx = find_header(self.STUDENT_FIRST_HEADERS)
            student_last_idx = find_header(self.STUDENT_LAST_HEADERS)
            class_idx = find_header(self.CLASS_HEADERS)
            teacher_idx = find_header(self.TEACHER_HEADERS)
            advisor_idx = find_header(self.ADVISOR_HEADERS)
            
            # Valid if we have either full name OR (first + last)
            # Class is optional - if missing, we'll use sheet name as class
            has_name = student_full_idx is not None or (student_first_idx is not None and student_last_idx is not None)
            
            if has_name:
                # If we have both first and last, prefer those over full name
                if student_first_idx is not None and student_last_idx is not None:
                    student_full_idx = None
                
                # Calculate quality score: number of exact matches in standard headers
                quality_score = 0
                if student_full_idx is not None and headers_lower[student_full_idx] in ["name", "student name", "student", "full name"]:
                    quality_score += 2
                if teacher_idx is not None and headers_lower[teacher_idx] == "teacher":
                    quality_score += 2
                if class_idx is not None and headers_lower[class_idx] in ["class", "class id", "section"]:
                    quality_score += 1
                
                candidates.append((quality_score, row_idx, {
                    "student_full": student_full_idx,
                    "student_first": student_first_idx,
                    "student_last": student_last_idx,
                    "class": class_idx,
                    "teacher": teacher_idx,
                    "advisor": advisor_idx
                }))
        
        # Return the best candidate (highest quality score, or first if tied)
        if candidates:
            candidates.sort(key=lambda x: x[0], reverse=True)
            _, best_row_idx, best_idxs = candidates[0]
            return best_row_idx, best_idxs
        
        # If no headers found, try fixed-width format (ROCL style)
        # Assume: Column A=Name, Column B=Class, Column C=Teacher
        print("[ROCL] No headers found, assuming fixed-width format (A=Name, B=Class, C=Teacher)")
        return None, {
            "student_full": 0,
            "student_first": None,
            "student_last": None,
            "class": 1,
            "teacher": 2,
            "advisor": None
        }

    def _extract_metadata_from_title_rows(self, worksheet, header_row):
        """
        Extract class and teacher info from title/metadata rows above the header.
        
        Looks for patterns like:
        - "Class 101/ Teacher: Ms. Smith"
        - "Teacher: John Doe"
        - "Class: 5A"
        
        Returns: (class_value, teacher_value)
        """
        import re
        
        class_val = None
        teacher_val = None
        
        # Scan rows above the header (or first 10 rows if no header)
        max_row = header_row if header_row is not None else 10
        
        for row_idx in range(max_row):
            row = list(worksheet.iter_rows(min_row=row_idx + 1, max_row=row_idx + 1, values_only=True))[0]
            
            # Check first column for metadata text
            if not row or not row[0]:
                continue
            
            text = str(row[0]).strip()
            
            # Pattern: "Class 101/ Teacher: Ms. Falleen and Ms. Kasandra"
            # or "Class 101/ Teachers: Ms. Falleen and Ms. Kasandra"
            match = re.search(r'Class\s+(\d+[^/]*)/\s*Teachers?:\s*(.+)', text, re.IGNORECASE)
            if match:
                class_val = match.group(1).strip()
                teacher_val = match.group(2).strip()
                continue
            
            # Pattern: "Teacher: John Doe" or "Teachers: John Doe" or "Teacher - John Doe"
            match = re.search(r'Teachers?[:\-]\s*(.+)', text, re.IGNORECASE)
            if match:
                teacher_val = match.group(1).strip()
                continue
            
            # Pattern: "Class: 5A" or "Class - 5A"
            match = re.search(r'Class[:\-]\s*(.+)', text, re.IGNORECASE)
            if match:
                class_val = match.group(1).strip()
                continue
        
        return class_val, teacher_val

    def _extract_and_clean(self, worksheet):
        """Extract rows → cleaned_records list[dict]."""
        header_row, idxs = self._find_header_row_and_indices(worksheet)
        
        # Extract metadata from title rows (class and teacher info)
        metadata_class, metadata_teacher = self._extract_metadata_from_title_rows(worksheet, header_row)
        
        # Get sheet name for use as default class if no class column exists
        sheet_name = worksheet.title if hasattr(worksheet, 'title') else "Unknown Class"
        
        # Prefer metadata class over sheet name
        default_class = metadata_class if metadata_class else sheet_name
        default_teacher = metadata_teacher if metadata_teacher else ""
        
        # If header_row is None, we're using fixed-width format - start from first non-empty row
        if header_row is None:
            start_row = 1
            # Find first row with actual data (non-empty student name)
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=100, values_only=True), 1):
                if row[0] and str(row[0]).strip():  # First column has content
                    start_row = row_idx
                    break
            header_row = start_row - 1
            data_start_row = start_row
        else:
            # For header-based format, data starts after the header row
            # header_row is 0-indexed row, so add 2 to get openpyxl's 1-indexed row number
            data_start_row = header_row + 2
        
        cleaned_records = []

        for row in worksheet.iter_rows(min_row=data_start_row, values_only=True):
            vals = list(row)

            def get(field):
                idx = idxs.get(field)
                return vals[idx] if idx is not None and idx < len(vals) else None

            full = get("student_full")
            first = get("student_first")
            last = get("student_last")
            class_val = get("class")
            teacher = get("teacher")
            advisor = get("advisor")

            # Need at least a name
            if not (full or first or last):
                continue
            
            # Use metadata values as defaults if column values are missing
            if not class_val:
                class_val = default_class
            
            if not teacher:
                teacher = default_teacher

            if full:
                student_name = self._clean_name(full)
            else:
                # Combine First + Last
                student_name = " ".join(
                    str(x).strip() for x in (first, last) if x
                ).strip()
                student_name = self._clean_name(student_name)

            record = {
                "teacher": self._clean_text(teacher) if teacher else "",
                "student": student_name,
                "class": self._clean_text(class_val),
            }

            if advisor:
                record["advisor"] = self._clean_text(advisor)

            cleaned_records.append(record)

        return cleaned_records

    # -------------------- existing helpers (unchanged) -------------------- #
    def _clean_name(self, name):
        if not name:
            return ""
        cleaned = str(name).strip().title()
        cleaned = self._fix_name_capitalization(cleaned)
        return cleaned

    def _fix_name_capitalization(self, name):
        if name.startswith("Mc") and len(name) > 2:
            name = "Mc" + name[2].upper() + name[3:]
        if name.startswith("Mac") and len(name) > 3:
            name = "Mac" + name[3].upper() + name[4:]
        if "O'" in name or "o'" in name:
            parts = name.split("'")
            if len(parts) == 2 and parts[0].upper() == "O":
                name = "O'" + parts[1].capitalize()
        return name

    def _clean_text(self, text):
        if not text:
            return ""
        return str(text).strip()

    def _create_cleaned_workbook(self, records):
        wb = Workbook()
        ws = wb.active
        ws.title = "Cleaned Data"

        if self.include_advisor:
            ws.append(["Teacher", "Student Name", "Class", "Advisor"])
        else:
            ws.append(["Teacher", "Student Name", "Class"])

        for record in records:
            if self.include_advisor:
                ws.append([
                    record.get("teacher", ""),
                    record.get("student", ""),
                    record.get("class", ""),
                    record.get("advisor", ""),
                ])
            else:
                ws.append([
                    record.get("teacher", ""),
                    record.get("student", ""),
                    record.get("class", ""),
                ])

        return wb

    def get_stage_name(self) -> str:
        return "Data Cleaning (ROCL Format)"


# ============================================================================
# PIPELINE FACTORIES
# ============================================================================

def get_cleaning_pipeline(school_format: str) -> ProcessingPipeline:
    """
    Get appropriate pipeline based on school format name.
    
    Args:
        school_format: School name or format (e.g., "ROCL", "ROCL with advisor", "default")
    
    Returns:
        ProcessingPipeline with appropriate cleaning stage + class splitting
    """
    school_format = school_format.strip().lower()
    
    if school_format == "rocl":
        return ProcessingPipeline(stages=[
            ROCLCleaningStage(include_advisor=False),
            ClassSplittingStage()
        ])
    elif school_format == "rocl with advisor":
        return ProcessingPipeline(stages=[
            ROCLCleaningStage(include_advisor=True),
            ClassSplittingStage()
        ])
    else:
        # Default: minimal cleaning + class splitting
        return ProcessingPipeline(stages=[
            DataCleaningStage(),
            ClassSplittingStage()
        ])


def get_cleaning_only_pipeline(school_format: str) -> ProcessingPipeline:
    """
    Get a pipeline that ONLY cleans data, does NOT split by class.
    Used for Stage 1 (Data Cleaning & Standardization).
    
    Args:
        school_format: School name or format
    
    Returns:
        ProcessingPipeline with only the cleaning stage
    """
    school_format = school_format.strip().lower()
    
    if school_format == "rocl":
        return ProcessingPipeline(stages=[ROCLCleaningStage(include_advisor=False)])
    elif school_format == "rocl with advisor":
        return ProcessingPipeline(stages=[ROCLCleaningStage(include_advisor=True)])
    elif school_format == "json_universal":
        return ProcessingPipeline(stages=[JSONUniversalCleaningStage()])
    else:
        return ProcessingPipeline(stages=[DataCleaningStage()])


class JSONUniversalCleaningStage(DataCleaningStage):
    """
    JSON-First Universal Mode Cleaning Stage
    
    Handles diverse Excel file formats with intelligent detection:
    - Grid/matrix layouts (classes in columns)
    - Schedule sheets (class-to-teacher mappings)
    - Standard row-based layouts
    - Metadata extraction from title rows
    - Teacher data verification to prevent overwrites
    
    This stage replaces the complex logic previously embedded in the UI.
    """
    
    def __init__(self):
        super().__init__()
    
    def process(self, data: dict) -> dict:
        """
        Process workbook using JSON-First Universal mode.
        
        Steps:
        1. Extract all sheets to JSON with format detection
        2. Normalize and match fields intelligently
        3. Create standardized 3-column output
        4. Export JSON for inspection
        """
        wb_in = data.get("workbook")
        if not wb_in:
            raise ValueError("Input workbook not found")
        
        print("\n[JSON-First Universal Mode]")
        print("="*60)
        
        # Step 1: Extract all sheets with smart format detection
        extracted_json = self._extract_all_sheets_to_json(wb_in)
        
        print(f"\n[Extraction Summary]")
        print(f"  Total sheets: {extracted_json['metadata']['total_sheets']}")
        print(f"  Processed: {len(extracted_json['metadata']['processed_sheets'])}")
        print(f"  Skipped: {len(extracted_json['metadata']['skipped_sheets'])}")
        
        # Step 2: Normalize data with intelligent field matching
        normalized_records = self._normalize_and_match_fields(extracted_json)
        
        print(f"\n[Normalization]")
        print(f"  Total records extracted: {len(normalized_records)}")
        
        if not normalized_records:
            print("  ⚠️ Warning: No student records found!")
            data["workbook"] = None
            data["cleaned_record_count"] = 0
            return data
        
        # Step 3: Create standardized output workbook (3 columns for Stage 2)
        wb_out = self._create_standardized_workbook(normalized_records)
        
        # Step 4: Export JSON for inspection (if path is set)
        if hasattr(self, 'json_output_path') and self.json_output_path:
            self._export_json(extracted_json, normalized_records)
        
        print("="*60)
        
        # Update data for next stage
        data["workbook"] = wb_out
        data["worksheet"] = wb_out.active
        data["extracted_json"] = extracted_json
        data["normalized_records"] = normalized_records
        data["cleaned_record_count"] = len(normalized_records)
        
        return data
    
    def _extract_all_sheets_to_json(self, workbook) -> dict:
        """
        Extract data from all sheets in workbook with format detection.
        
        Returns:
            {
                'metadata': {
                    'total_sheets': int,
                    'processed_sheets': [sheet_names],
                    'skipped_sheets': [sheet_names]
                },
                'raw_data': {
                    'sheet_name': {
                        'headers': [header_row],
                        'rows': [[row_data]]
                    }
                }
            }
        """
        json_structure = {
            'metadata': {
                'total_sheets': len(workbook.sheetnames),
                'processed_sheets': [],
                'skipped_sheets': []
            },
            'raw_data': {}
        }
        
        # Process each sheet
        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            
            # Skip obviously non-data sheets
            if self._should_skip_sheet(sheet_name):
                json_structure['metadata']['skipped_sheets'].append(sheet_name)
                print(f"  Skipping sheet: {sheet_name}")
                continue
            
            print(f"  Processing sheet: {sheet_name}")
            
            # Extract sheet data
            sheet_data = self._extract_sheet_data(ws)
            # Add sheets with data OR schedule sheets (which have metadata but no rows)
            if sheet_data['rows'] or sheet_data['metadata'].get('is_schedule'):
                json_structure['raw_data'][sheet_name] = sheet_data
                json_structure['metadata']['processed_sheets'].append(sheet_name)
            else:
                json_structure['metadata']['skipped_sheets'].append(sheet_name)
        
        return json_structure
    
    def _should_skip_sheet(self, sheet_name: str) -> bool:
        """Determine if sheet should be skipped (non-data sheets)"""
        skip_keywords = [
            'instructions', 'notes', 'readme', 'info', 'help',
            'template', 'example', 'summary', 'cover'
        ]
        sheet_lower = sheet_name.lower().strip()
        return any(keyword in sheet_lower for keyword in skip_keywords)
    
    def _extract_sheet_data(self, worksheet) -> dict:
        """
        Extract data from single sheet.
        
        Returns:
            {
                'headers': [header_candidates],
                'rows': [[row_values]],
                'metadata': {
                    'total_rows': int,
                    'data_start_row': int,
                    'header_row': int,
                    'class_from_title': str (optional),
                    'teacher_from_title': str (optional),
                    'is_schedule': bool (if this is a schedule/mapping sheet)
                }
            }
        """
        # Check if this is a schedule sheet (class-to-teacher mapping)
        if self._is_schedule_sheet(worksheet):
            print(f"  [Schedule Sheet] Detected - extracting teacher mappings...")
            return self._extract_schedule_data(worksheet)
        
        # Check if this is a grid/matrix format (classes in columns, students in rows)
        if self._is_grid_format(worksheet):
            print(f"  [Grid Format] Detected grid layout - transforming data...")
            return self._extract_grid_format_data(worksheet)
        
        # First, extract metadata from title rows (before looking for headers)
        metadata_class, metadata_teacher = self._extract_metadata_from_title_rows(worksheet)
        
        # Find potential header row (look in first 10 rows)
        header_row_idx = None
        headers = []
        
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=10, values_only=True)):
            if self._looks_like_header_row(row):
                header_row_idx = row_idx + 1  # 1-indexed
                headers = [self._clean_header(cell) for cell in row]
                break
        
        # If no headers found, use first non-empty row
        if header_row_idx is None:
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=20, values_only=True)):
                if any(cell for cell in row):
                    header_row_idx = row_idx + 1
                    headers = [f"Column_{i}" for i in range(len(row))]
                    break
        
        # Extract data rows
        data_rows = []
        start_row = (header_row_idx + 1) if header_row_idx else 1
        
        for row in worksheet.iter_rows(min_row=start_row, values_only=True):
            # Skip completely empty rows
            if not any(cell for cell in row):
                continue
            
            # Skip metadata/footer rows (Teacher:, Advisor:, etc.)
            if row[0] and self._is_metadata_row(str(row[0])):
                continue
            
            # Convert row to list and clean values
            cleaned_row = [self._clean_cell_value(cell) for cell in row]
            data_rows.append(cleaned_row)
        
        return {
            'headers': headers,
            'rows': data_rows,
            'metadata': {
                'total_rows': len(data_rows),
                'data_start_row': start_row,
                'header_row': header_row_idx,
                'class_from_title': metadata_class,
                'teacher_from_title': metadata_teacher
            }
        }
    
    def _is_schedule_sheet(self, worksheet) -> bool:
        """
        Detect if this is a schedule sheet (class-to-teacher mapping).
        
        Schedule sheets typically have:
        - Headers like "Class", "Teacher", "Time", etc.
        - Relatively few rows (< 20)
        - Class codes/IDs in one column and teacher names in another
        """
        # Check first 10 rows for schedule indicators
        headers_found = []
        for row_idx in range(1, min(11, worksheet.max_row + 1)):
            row = list(worksheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
            for cell in row:
                if cell:
                    headers_found.append(str(cell).lower().strip())
        
        # Look for schedule-specific headers
        has_class_col = any('class' in h for h in headers_found)
        has_teacher_col = any('teacher' in h for h in headers_found)
        has_time_col = any(word in h for h in headers_found for word in ['time', 'schedule', 'transition', 'photo'])
        
        # Schedule sheets are usually short (not student rosters)
        is_short = worksheet.max_row < 20
        
        return has_class_col and has_teacher_col and (has_time_col or is_short)
    
    def _extract_schedule_data(self, worksheet) -> dict:
        """
        Extract class-to-teacher mapping from schedule sheet.
        
        Returns a special format indicating this is a schedule sheet.
        """
        # Find header row
        header_row_idx = None
        headers = []
        
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=10, values_only=True)):
            row_text = [str(cell).lower().strip() if cell else "" for cell in row]
            if 'class' in row_text and 'teacher' in row_text:
                header_row_idx = row_idx + 1
                headers = [str(cell).strip() if cell else "" for cell in row]
                break
        
        if not header_row_idx:
            return {'headers': [], 'rows': [], 'metadata': {'is_schedule': False}}
        
        # Find class and teacher column indices
        class_idx = None
        teacher_idx = None
        for idx, header in enumerate(headers):
            header_lower = header.lower()
            if 'class' in header_lower and class_idx is None:
                class_idx = idx
            if 'teacher' in header_lower and teacher_idx is None:
                teacher_idx = idx
        
        if class_idx is None or teacher_idx is None:
            return {'headers': [], 'rows': [], 'metadata': {'is_schedule': False}}
        
        # Extract class-teacher mappings
        schedule_data = []
        for row in worksheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if not row or len(row) <= max(class_idx, teacher_idx):
                continue
            
            class_val = row[class_idx]
            teacher_val = row[teacher_idx]
            
            if not class_val or not teacher_val:
                continue
            
            class_str = str(class_val).strip()
            teacher_str = str(teacher_val).strip()
            
            # Skip empty or header-like rows
            if not class_str or not teacher_str or class_str.lower() == 'class':
                continue
            
            schedule_data.append({
                'class': class_str,
                'teacher': teacher_str
            })
        
        return {
            'headers': headers,
            'rows': [],  # Don't include raw rows for schedule sheets
            'metadata': {
                'is_schedule': True,
                'schedule_data': schedule_data,
                'total_rows': len(schedule_data),
                'data_start_row': header_row_idx + 1,
                'header_row': header_row_idx
            }
        }
    
    def _is_grid_format(self, worksheet) -> bool:
        """
        Detect if worksheet uses grid/matrix format (classes in columns).
        
        Grid format characteristics:
        - Row 1 has multiple class names across columns (Class A, Class B, etc.)
        - Rows have labels like "Teacher", "Student 1", "Student 2", etc. in column A
        - Each column represents a different class
        """
        # Check first row for multiple class indicators
        first_row = list(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        
        # Count how many cells in row 1 contain "class" or grade level indicators
        class_indicators = 0
        for cell in first_row[1:6]:  # Check first 5 columns after column A
            if not cell:
                continue
            cell_str = str(cell).lower()
            if 'class' in cell_str or 'grade' in cell_str or any(level in cell_str for level in ['3k', '4k', 'prek', 'pre-k', 'k']):
                class_indicators += 1
        
        # If multiple columns have class indicators, likely grid format
        if class_indicators >= 2:
            return True
        
        # Also check if column A has row labels like "Teacher", "Student 1", etc.
        row_labels = []
        for row_idx in range(2, min(10, worksheet.max_row + 1)):
            row = list(worksheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
            if row[0]:
                row_labels.append(str(row[0]).lower().strip())
        
        # Check if we have typical grid format labels
        has_teacher_row = any('teacher' in label for label in row_labels)
        has_student_rows = sum(1 for label in row_labels if 'student' in label) >= 3
        
        return has_teacher_row and has_student_rows
    
    def _extract_grid_format_data(self, worksheet) -> dict:
        """
        Extract data from grid/matrix format where classes are in columns.
        
        Example format:
        Row 1: | Column1      | Class A 3k | Class B Prek | Class C 3k |
        Row 2: | Teacher      | Ms Martha  | Ms Nicolle   | Ms Tafia   |
        Row 3: | Student 1    | John Doe   | Jane Smith   | Bob Lee    |
        
        Transforms to standard row-per-student format.
        """
        # Get all rows
        all_rows = list(worksheet.iter_rows(min_row=1, values_only=True))
        
        if len(all_rows) < 3:
            return {'headers': [], 'rows': [], 'metadata': {'total_rows': 0, 'data_start_row': 1, 'header_row': None}}
        
        # Row 1 contains class names
        header_row = all_rows[0]
        class_names = [str(cell).strip() if cell else f"Column_{i}" for i, cell in enumerate(header_row)]
        
        # Find teacher row(s) - rows where column A contains "teacher"
        teacher_data = {}
        student_rows = []
        
        for row_idx, row in enumerate(all_rows[1:], 2):  # Start from row 2
            row_label = str(row[0]).lower().strip() if row[0] else ""
            
            if 'teacher' in row_label:
                # This is a teacher row - extract teachers for each class
                for col_idx, cell in enumerate(row[1:], 1):  # Skip column A
                    if cell and col_idx < len(class_names):
                        class_name = class_names[col_idx]
                        if class_name not in teacher_data:
                            teacher_data[class_name] = []
                        teacher_data[class_name].append(str(cell).strip())
            
            elif 'student' in row_label:
                # This is a student row
                student_rows.append(row)
        
        # Transform: one row per student
        transformed_rows = []
        for student_row in student_rows:
            for col_idx, student_name in enumerate(student_row[1:], 1):  # Skip column A (label)
                if not student_name or col_idx >= len(class_names):
                    continue
                
                student_name = str(student_name).strip()
                if not student_name:
                    continue
                
                class_name = class_names[col_idx]
                teachers = ' / '.join(teacher_data.get(class_name, []))
                
                # Create row: [Teacher, First Name, Last Name, Class]
                # For now, use full name and empty first/last
                transformed_rows.append([teachers, student_name, '', class_name])
        
        # Return in standard format
        return {
            'headers': ['Teacher', 'Student Name', 'Student First', 'Class'],
            'rows': transformed_rows,
            'metadata': {
                'total_rows': len(transformed_rows),
                'data_start_row': 2,
                'header_row': 1,
                'class_from_title': None,
                'teacher_from_title': None
            }
        }
    
    def _looks_like_header_row(self, row) -> bool:
        """Determine if row looks like headers"""
        if not row:
            return False
        
        # Count non-empty cells
        non_empty_count = sum(1 for cell in row if cell is not None and str(cell).strip())
        
        # Skip if only 1 cell has content (likely a title row like "CLASS LIST")
        if non_empty_count < 2:
            return False
        
        # Skip common title patterns
        first_cell = str(row[0]).strip().upper() if row[0] else ""
        title_patterns = ['CLASS LIST', 'ROSTER', 'STUDENT LIST', 'SITE:', 'YEAR:']
        if any(first_cell.startswith(pattern) for pattern in title_patterns):
            return False
        
        # Count text cells vs numeric cells
        text_count = 0
        total_count = 0
        
        for cell in row:
            if cell is not None:
                total_count += 1
                if isinstance(cell, str) and len(str(cell).strip()) > 0:
                    text_count += 1
        
        # Header row should be mostly text and have multiple columns
        return total_count >= 2 and (text_count / total_count) >= 0.7
    
    def _clean_header(self, header) -> str:
        """Clean header name"""
        if header is None:
            return ""
        return str(header).strip()
    
    def _clean_cell_value(self, cell):
        """Clean individual cell value"""
        if cell is None:
            return ""
        if isinstance(cell, str):
            return cell.strip()
        return cell
    
    def _is_metadata_row(self, text: str) -> bool:
        """
        Check if a row is metadata/footer (not student data).
        
        Examples of metadata rows to skip:
        - "Teacher: Clever Caterpillars /Ms. Falleen and Ms. Kasandra"
        - "Advisor:"
        - "CLASS LIST"
        - "Site: Skillman..."
        - "Year: 2025-26"
        """
        text_lower = text.lower().strip()
        
        # Common metadata patterns
        metadata_patterns = [
            'teacher:', 'advisor:', 'class list', 'site:', 'year:', 
            'ed director', 'family worker', 'total', 'count',
            'class ' + text_lower.split()[0] if text_lower.startswith('class') else ''
        ]
        
        # Check if starts with any metadata pattern
        for pattern in metadata_patterns:
            if pattern and text_lower.startswith(pattern):
                return True
        
        # Check if it's just numbers (like row numbers "#")
        if text.strip() in ['#', '']:
            return True
        
        return False
    
    def _extract_metadata_from_title_rows(self, worksheet):
        """
        Extract class and teacher info from metadata/title rows.
        
        Common patterns:
        - "Class 101/ Teacher: Ms. Falleen and Ms. Kasandra"
        - "Class 101 / Teacher: Clever Caterpillars /Ms. Falleen and Ms. Kasandra"
        - "Teacher: John Doe"
        - "Class: 5A"
        
        Returns: (class_value, teacher_value)
        """
        import re
        
        class_val = None
        teacher_val = None
        
        # Scan first 10 rows for metadata patterns
        for row_idx in range(10):
            row = list(worksheet.iter_rows(min_row=row_idx + 1, max_row=row_idx + 1, values_only=True))[0]
            
            if not row or not row[0]:
                continue
            
            text = str(row[0]).strip()
            
            # Pattern: "Class 101/ Teacher: Ms. Falleen and Ms. Kasandra"
            # or "Class 101 / Teachers: Clever Caterpillars /Ms. Falleen and Ms. Kasandra"
            match = re.search(r'Class\s+(\d+[^/]*)/\s*Teachers?:\s*(.+)', text, re.IGNORECASE)
            if match:
                class_val = match.group(1).strip()
                teacher_val = match.group(2).strip()
                print(f"  [Metadata] Found class: {class_val}, teacher: {teacher_val}")
                continue
            
            # Pattern: "Teacher: John Doe" or "Teachers: John Doe" or "Teacher - John Doe"
            match = re.search(r'Teachers?[:\-]\s*(.+)', text, re.IGNORECASE)
            if match:
                teacher_val = match.group(1).strip()
                print(f"  [Metadata] Found teacher: {teacher_val}")
                continue
            
            # Pattern: "Class: 5A" or "Class - 5A"
            match = re.search(r'Class[:\-]\s*(.+)', text, re.IGNORECASE)
            if match:
                class_val = match.group(1).strip()
                print(f"  [Metadata] Found class: {class_val}")
                continue
        
        return class_val, teacher_val
    
    def _normalize_and_match_fields(self, extracted_json: dict) -> list:
        """
        Smart field matching: Match extracted data to standardized fields.
        
        Target fields:
        - student_name (or first_name + last_name)
        - class_id / class_name
        - teacher_name
        - grade (optional)
        - advisor (optional)
        
        Returns list of normalized records
        """
        normalized_records = []
        
        # First, scan to see if student sheets already have teacher data
        has_teacher_column = self._check_for_teacher_column(extracted_json)
        
        # Extract teacher mapping from schedule sheets (only if needed)
        teacher_map = {}
        if not has_teacher_column:
            teacher_map = self._build_teacher_map(extracted_json)
            if teacher_map:
                print(f"  [Schedule Mapping] No teacher column found in student data")
                print(f"  [Schedule Mapping] Using Schedule sheet with {len(teacher_map)} class-to-teacher mappings")
        else:
            print(f"  [Teacher Column] Found teacher data in student sheets - using it directly")
        
        # Field matching patterns
        field_patterns = {
            'student_name': [
                'name', 'student', 'student name', 'full name', 
                'student full name', 'stu name'
            ],
            'first_name': [
                'first name', 'fname', 'first', 'given name',
                'student first name', 'given'
            ],
            'last_name': [
                'last name', 'lname', 'last', 'surname',
                'student last name', 'family name'
            ],
            'class_id': [
                'class', 'class id', 'class section', 'section',
                'class section id', 'ocl', 'grade', 'homeroom'
            ],
            'teacher': [
                'teacher', 'teacher name', 'homeroom teacher',
                'instructor', 'tchr', 'homeroom'
            ],
            'grade': [
                'grade', 'grade level', 'year', 'level'
            ],
            'advisor': [
                'advisor', 'adviser', 'adv'
            ]
        }
        
        # Process each sheet's data
        for sheet_name, sheet_data in extracted_json['raw_data'].items():
            headers = sheet_data['headers']
            rows = sheet_data['rows']
            metadata = sheet_data.get('metadata', {})
            
            # Skip schedule sheets (they don't contain student data)
            if metadata.get('is_schedule'):
                continue
            
            # Get metadata values (class and teacher from title rows)
            default_class = metadata.get('class_from_title') or sheet_name
            default_teacher = metadata.get('teacher_from_title', '')
            
            # Match headers to fields
            field_indices = self._match_headers_to_fields(headers, field_patterns)
            
            # Process each row
            for row in rows:
                record = self._create_normalized_record(
                    row, field_indices, sheet_name, 
                    default_class=default_class,
                    default_teacher=default_teacher
                )
                if record and record.get('student_name'):  # Must have student name
                    # Try to fill in teacher from schedule map if missing or empty
                    if (not record.get('teacher') or not record.get('teacher').strip()) and teacher_map:
                        grade_or_class = record.get('grade') or record.get('class_id', '')
                        record['teacher'] = self._lookup_teacher(grade_or_class, teacher_map)
                    
                    normalized_records.append(record)
        
        return normalized_records
    
    def _check_for_teacher_column(self, extracted_json: dict) -> bool:
        """
        Scan student sheets to verify if they contain teacher data.
        
        Returns True if at least one student sheet has a teacher column with data.
        """
        for sheet_name, sheet_data in extracted_json['raw_data'].items():
            metadata = sheet_data.get('metadata', {})
            
            # Skip schedule sheets and grid format sheets (which extract teachers separately)
            if metadata.get('is_schedule'):
                continue
            
            headers = sheet_data['headers']
            rows = sheet_data['rows']
            
            if not headers or not rows:
                continue
            
            # Look for teacher column in headers
            teacher_col_idx = None
            for idx, header in enumerate(headers):
                header_lower = str(header).lower().strip() if header else ""
                if 'teacher' in header_lower:
                    teacher_col_idx = idx
                    break
            
            # If found teacher column, check if it has actual data
            if teacher_col_idx is not None:
                # Check first 10 rows for non-empty teacher data
                has_data = False
                for row in rows[:10]:
                    if teacher_col_idx < len(row):
                        teacher_val = row[teacher_col_idx]
                        if teacher_val and str(teacher_val).strip():
                            has_data = True
                            break
                
                if has_data:
                    return True
        
        return False
    
    def _build_teacher_map(self, extracted_json: dict) -> dict:
        """
        Build a class-to-teacher mapping from schedule sheets.
        
        Returns: {class_code: teacher_name}
        """
        teacher_map = {}
        
        for sheet_name, sheet_data in extracted_json['raw_data'].items():
            metadata = sheet_data.get('metadata', {})
            
            if metadata.get('is_schedule') and metadata.get('schedule_data'):
                for mapping in metadata['schedule_data']:
                    class_code = mapping['class']
                    teacher_name = mapping['teacher']
                    teacher_map[class_code] = teacher_name
                    
                    # Also add normalized versions for matching
                    # "8.3 / 8.1" should match "8.3" and "8.1"
                    if '/' in class_code:
                        for part in class_code.split('/'):
                            part = part.strip()
                            if part:
                                teacher_map[part] = teacher_name
        
        return teacher_map
    
    def _lookup_teacher(self, grade_level: str, teacher_map: dict) -> str:
        """
        Lookup teacher from grade level using the teacher map.
        
        Grade level examples: "6th Grade", "7th Grade", "8th Grade"
        Class codes examples: "6.1", "7.2", "8.3"
        """
        if not grade_level or not teacher_map:
            return ''
        
        grade_str = str(grade_level).strip()
        
        # Try exact match first
        if grade_str in teacher_map:
            return teacher_map[grade_str]
        
        # Extract grade number (e.g., "6th Grade" → "6", "7th Grade" → "7")
        import re
        grade_match = re.search(r'(\d+)', grade_str)
        if not grade_match:
            return ''
        
        grade_num = grade_match.group(1)
        
        # Look for classes starting with this grade number
        # Collect all teachers for this grade
        teachers_for_grade = []
        for class_code, teacher in teacher_map.items():
            # Match "6.1", "6.2", etc. for grade 6
            if class_code.startswith(f"{grade_num}.") or class_code == grade_num:
                if teacher not in teachers_for_grade:
                    teachers_for_grade.append(teacher)
        
        # Return combined list of all teachers for this grade
        if teachers_for_grade:
            return ' / '.join(teachers_for_grade)
        
        return ''
    
    def _match_headers_to_fields(self, headers: list, field_patterns: dict) -> dict:
        """
        Match headers to standardized field names using pattern matching.
        Prioritizes more specific patterns (first_name, last_name) over generic ones (student_name).
        
        Returns: {field_name: column_index}
        """
        field_indices = {}
        
        # Match in priority order: more specific first, then generic
        priority_order = ['first_name', 'last_name', 'student_name', 'class_id', 'teacher', 'grade', 'advisor']
        
        for field_name in priority_order:
            if field_name not in field_patterns:
                continue
            patterns = field_patterns[field_name]
            for idx, header in enumerate(headers):
                # Skip if this index is already used by a higher priority field
                if idx in field_indices.values():
                    continue
                if self._header_matches_patterns(header, patterns):
                    field_indices[field_name] = idx
                    break
        
        return field_indices
    
    def _header_matches_patterns(self, header: str, patterns: list) -> bool:
        """Check if header matches any of the patterns"""
        if not header:
            return False
        
        header_clean = header.lower().strip()
        
        # Exact match first
        for pattern in patterns:
            if header_clean == pattern.lower():
                return True
        
        # Substring match second
        for pattern in patterns:
            if pattern.lower() in header_clean:
                return True
        
        return False
    
    def _create_normalized_record(self, row: list, field_indices: dict, sheet_name: str, 
                                  default_class: str = None, default_teacher: str = None) -> dict:
        """Create normalized record from row data"""
        def get_field(field_name):
            idx = field_indices.get(field_name)
            if idx is not None and idx < len(row):
                return self._clean_text(row[idx])
            return ""
        
        # Get student name (prefer full name, fallback to first+last)
        student_name = get_field('student_name')
        if not student_name:
            first_name = get_field('first_name')
            last_name = get_field('last_name')
            if first_name or last_name:
                student_name = f"{first_name} {last_name}".strip()
        
        if not student_name:
            return None  # Skip records without student name
        
        # Get class and teacher (use metadata defaults if not in columns)
        class_id = get_field('class_id') or default_class or sheet_name
        teacher = get_field('teacher') or default_teacher or ''
        
        # Create normalized record
        record = {
            'student_name': self._clean_name(student_name),
            'class_id': class_id,
            'teacher': teacher,
            'grade': get_field('grade'),
            'advisor': get_field('advisor'),
            'source_sheet': sheet_name
        }
        
        return record
    
    def _clean_text(self, text) -> str:
        """Clean text field"""
        if not text:
            return ""
        return str(text).strip()
    
    def _clean_name(self, name: str) -> str:
        """Clean and format names properly"""
        if not name:
            return ""
        
        # Basic cleaning
        cleaned = str(name).strip()
        
        # Title case if all caps
        if cleaned.isupper():
            cleaned = cleaned.title()
        
        # Fix common name patterns
        cleaned = self._fix_name_capitalization(cleaned)
        
        return cleaned
    
    def _fix_name_capitalization(self, name: str) -> str:
        """Fix capitalization for names with prefixes"""
        if name.startswith("Mc") and len(name) > 2:
            name = "Mc" + name[2].upper() + name[3:]
        elif name.startswith("Mac") and len(name) > 3:
            name = "Mac" + name[3].upper() + name[4:]
        elif "O'" in name or "o'" in name:
            parts = name.split("'")
            if len(parts) == 2 and parts[0].upper() == "O":
                name = "O'" + parts[1].capitalize()
        return name
    
    def _create_standardized_workbook(self, normalized_records: list):
        """Create standardized Excel output from normalized records"""
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Standardized Data"
        
        # Standard headers - simplified to match ROCL output for Stage 2 compatibility
        ws.append(["Teacher", "Student Name", "Class"])
        
        # Add records
        for record in normalized_records:
            ws.append([
                record.get('teacher', ''),
                record.get('student_name', ''),
                record.get('class_id', '')
            ])
        
        return wb
    
    def _export_json(self, extracted_json: dict, normalized_records: list):
        """Export JSON for inspection"""
        import json
        from pathlib import Path
        from datetime import datetime, date, time
        
        class DateTimeEncoder(json.JSONEncoder):
            """Custom JSON encoder to handle datetime/date/time objects"""
            def default(self, obj):
                if isinstance(obj, (datetime, date)):
                    return obj.isoformat()
                elif isinstance(obj, time):
                    return obj.isoformat()
                return super().default(obj)
        
        output_data = {
            'extraction_summary': extracted_json['metadata'],
            'raw_data': extracted_json['raw_data'],
            'normalized_records': normalized_records,
            'record_count': len(normalized_records)
        }
        
        json_path = Path(self.json_output_path)
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(output_data, f, indent=2, ensure_ascii=False, cls=DateTimeEncoder)
        
        print(f"  JSON exported to: {json_path}")
    
    def get_stage_name(self) -> str:
        return "JSON-First Universal Processing"

