import re
import sys
from pathlib import Path
from abc import ABC, abstractmethod
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QLineEdit, QFileDialog, QTextEdit,
    QMessageBox, QGroupBox, QComboBox, QCheckBox
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from PyQt6.QtGui import QFont
from openpyxl import load_workbook, Workbook


# ============================================================================
# PIPELINE STAGES - Base classes for extensible processing pipeline
# ============================================================================

class PipelineStage(ABC):
    """Base class for data processing stages"""
    
    @abstractmethod
    def process(self, data: dict) -> dict:
        """
        Process data and return modified data.
        data should contain: {"workbook": wb, "worksheet": ws, "headers": list}
        """
        pass
    
    @abstractmethod
    def get_stage_name(self) -> str:
        """Return human-readable name of this stage"""
        pass


class DataCleaningStage(PipelineStage):
    """
    Minimal base cleaning stage (for backwards compatibility).
    
    For JSON-First Universal mode, use JSONUniversalCleaningStage from cleaning_stages.py.
    This base class is kept minimal to avoid UI file bloat.
    """
    
    def __init__(self, export_json=False, json_output_path=None):
        """
        Initialize data cleaning stage.
        
        Args:
            export_json: If True, export intermediate JSON for inspection
            json_output_path: Path to save JSON file (optional)
        """
        self.export_json = export_json
        self.json_output_path = json_output_path
    
    def process(self, data: dict) -> dict:
        """
        Minimal processing - just pass through.
        Override in subclasses for actual functionality.
        """
        wb_in = data.get("workbook")
        if not wb_in:
            raise ValueError("Input workbook not found")
        
        # Simple pass-through for backwards compatibility
        data["cleaned_record_count"] = 0
        return data
    
    def get_stage_name(self) -> str:
        return "Basic Data Cleaning"


class ClassSplittingStage(PipelineStage):
    """Stage 2: Group students by class and generate separate sheets"""
    
    def __init__(self, header_mappings: dict = None):
        """
        Initialize with optional custom header mappings.
        
        Args:
            header_mappings: Dict with keys 'class', 'teacher', 'student'
                           and list of possible column names as values
        """
        self.header_mappings = header_mappings or self._get_default_mappings()
    
    @staticmethod
    def _get_default_mappings():
        return {
            "class": ["CLASS", "Class", "Class Section ID", "Class Section", "Class ID"],
            "teacher": ["Teacher", "TEACHER"],
            "student": [
                "Name",
                "Student",
                "STUDENT",
                "Student Name",
                "StudentName",
                "Student Name(s)"
            ]
        }
    
    def process(self, data: dict) -> dict:
        """Group students by class and create output workbook"""
        ws_in = data.get("worksheet")
        if not ws_in:
            raise ValueError("Input worksheet not found in data")
        
        header_row = [cell.value for cell in next(ws_in.iter_rows(min_row=1, max_row=1))]
        
        # Find column indices with error handling
        try:
            class_idx = self._find_column_index(header_row, self.header_mappings["class"])
            teacher_idx = self._find_column_index(header_row, self.header_mappings["teacher"])
            student_idx = self._find_column_index(header_row, self.header_mappings["student"])
        except ValueError as e:
            raise ValueError(f"Column mapping error: {str(e)}")
        
        # Group students by class
        groups = self._group_by_class(ws_in, class_idx, teacher_idx, student_idx)
        
        # Create output workbook with separate sheets per class
        wb_out = self._create_output_workbook(groups)
        
        data["output_workbook"] = wb_out
        data["class_groups"] = groups
        data["row_count"] = sum(len(info["students"]) for info in groups.values())
        
        return data
    
    def _find_column_index(self, header_row, candidates):
        """Find column index by matching against candidate names"""
        normalized = {}
        for i, value in enumerate(header_row):
            key = self._normalize_header(value)
            if key:
                normalized[key] = i
        for cand in candidates:
            key = self._normalize_header(cand)
            if key in normalized:
                return normalized[key]
        raise ValueError(f"Could not find any of headers: {candidates}")

    @staticmethod
    def _normalize_header(value):
        if value is None:
            return ""
        normalized = re.sub(r"[^a-z0-9]+", " ", str(value).strip().lower())
        return normalized.strip()
    
    def _group_by_class(self, worksheet, class_idx, teacher_idx, student_idx):
        """Group students by class, removing duplicates"""
        groups = {}
        
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            
            cls_val = row[class_idx]
            student_val = row[student_idx]
            teacher_val = row[teacher_idx] if teacher_idx is not None else None
            
            if not cls_val or not student_val:
                continue
            
            cls = str(cls_val).strip()
            student = str(student_val).strip()
            teacher = str(teacher_val).strip() if teacher_val else ""
            
            if cls not in groups:
                groups[cls] = {"teacher": teacher, "students": set()}
            
            if teacher and not groups[cls]["teacher"]:
                groups[cls]["teacher"] = teacher
            
            # Use set to avoid duplicates
            groups[cls]["students"].add(student)
        
        for cls in groups:
            groups[cls]["students"] = sorted(groups[cls]["students"])
        return groups
    
    def _create_output_workbook(self, groups):
        """Create Excel workbook with separate sheets per class"""
        wb_out = Workbook()
        first_sheet = True
        
        for cls in sorted(groups.keys()):
            info = groups[cls]
            title = self._make_sheet_name(cls)
            
            if first_sheet:
                ws = wb_out.active
                ws.title = title
                first_sheet = False
            else:
                ws = wb_out.create_sheet(title)
            
            # Headers
            ws["A1"] = "Student Name"
            ws["B1"] = "Class"
            
            row_idx = 2
            for student in info["students"]:
                ws.cell(row=row_idx, column=1, value=student)
                ws.cell(row=row_idx, column=2, value=cls)
                row_idx += 1
            
            # Footer info
            teacher_footer = f"Teacher: {info['teacher']}" if info["teacher"] else "Teacher:"
            ws.cell(row=row_idx, column=1, value=teacher_footer)
            row_idx += 1
            ws.cell(row=row_idx, column=1, value="Advisor:")
        
        return wb_out
    
    def _make_sheet_name(self, class_value):
        """Create valid Excel sheet name by replacing invalid characters"""
        cls_str = str(class_value)
        # Excel sheet names cannot contain: \ / * ? : [ ]
        cleaned = re.sub(r'[\\/*?:\[\]]', '-', cls_str)
        return ("Class " + cleaned)[:31]
    
    def get_stage_name(self) -> str:
        return "Class Splitting"


# ============================================================================
# PROCESSING PIPELINE - Orchestrates multiple stages
# ============================================================================

class ProcessingPipeline:
    """Orchestrates multiple data processing stages"""
    
    def __init__(self, stages: list = None):
        """Initialize with list of PipelineStage instances"""
        self.stages = stages or self._get_default_stages()
    
    @staticmethod
    def _get_default_stages():
        """Default pipeline: clean data, then split by class"""
        return [
            DataCleaningStage(),
            ClassSplittingStage()
        ]
    
    def execute(self, source_path: str, output_path: str, progress_callback=None):
        """
        Execute full processing pipeline.
        
        Args:
            source_path: Path to input Excel file
            output_path: Path to save output Excel file
            progress_callback: Optional callback function for progress updates
        
        Returns:
            tuple: (success: bool, message: str, result_data: dict)
        """
        try:
            data = {}
            
            # Load input file
            if progress_callback:
                progress_callback("Loading source file...")
            
            wb_in = load_workbook(source_path, data_only=True)
            ws_in = wb_in.active
            data["workbook"] = wb_in
            data["worksheet"] = ws_in
            
            # Execute each pipeline stage
            for stage in self.stages:
                if progress_callback:
                    progress_callback(f"Running: {stage.get_stage_name()}...")
                
                data = stage.process(data)
            
            # Save output workbook if it exists
            wb_out = data.get("output_workbook")
            if wb_out:
                if progress_callback:
                    progress_callback("Saving output file...")
                wb_out.save(output_path)
            else:
                # For cleaning-only pipelines, save whichever workbook the stages produced
                wb_to_save = data.get("workbook") or wb_in
                if progress_callback:
                    progress_callback("Saving cleaned file...")
                wb_to_save.save(output_path)
            
            # Generate result message
            if "class_groups" in data:
                # Pipeline includes ClassSplittingStage
                row_count = data.get("row_count", 0)
                class_count = len(data.get("class_groups", {}))
                message = f"Successfully processed {row_count} records into {class_count} class sheets!"
            else:
                # Pipeline is cleaning-only
                message = f"Successfully cleaned and standardized data!"
            
            return True, message, data
        
        except Exception as e:
            return False, f"Error: {str(e)}", {}


# ============================================================================
# UI THREAD - Runs pipeline asynchronously
# ============================================================================

class ProcessThread(QThread):
    """Thread for processing Excel file to avoid blocking UI"""
    progress = pyqtSignal(str)
    finished = pyqtSignal(bool, str)

    def __init__(self, source_path, output_path, pipeline: ProcessingPipeline = None):
        super().__init__()
        self.source_path = source_path
        self.output_path = output_path
        self.pipeline = pipeline or ProcessingPipeline()

    def run(self):
        success, message, _ = self.pipeline.execute(
            self.source_path,
            self.output_path,
            progress_callback=self.progress.emit
        )
        self.finished.emit(success, message)


# ============================================================================
# GUI - Main application interface with two-stage workflow
# ============================================================================

class ClassRosterGUI(QMainWindow):
    def __init__(self):
        super().__init__()
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("Class Roster Generator - Two Stage Pipeline")
        self.setMinimumSize(900, 700)

        # Central widget
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(15)
        main_layout.setContentsMargins(15, 15, 15, 15)

        # Title
        title = QLabel("Class Roster Generator")
        title_font = QFont()
        title_font.setPointSize(18)
        title_font.setBold(True)
        title.setFont(title_font)
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        main_layout.addWidget(title)

        # Separator line
        separator = QLabel("─" * 80)
        separator.setAlignment(Qt.AlignmentFlag.AlignCenter)
        main_layout.addWidget(separator)

        # ========================================================================
        # TWO-STAGE CONTAINER (Horizontal: Stage 1 and Stage 2 side by side)
        # ========================================================================
        two_stage_layout = QHBoxLayout()
        two_stage_layout.setSpacing(20)

        # ========================================================================
        # LEFT COLUMN: STAGE 1: DATA CLEANING
        # ========================================================================
        stage1_container = QVBoxLayout()

        stage1_label = QLabel("STAGE 1: Data Cleaning\n& Standardization")
        stage1_font = QFont()
        stage1_font.setPointSize(11)
        stage1_font.setBold(True)
        stage1_label.setFont(stage1_font)
        stage1_label.setStyleSheet("color: #2196F3;")
        stage1_container.addWidget(stage1_label)


        stage1_desc = QLabel(
            "Convert raw school data\n(various formats) into\none standardized Excel file"
        )
        stage1_desc.setStyleSheet("color: #666; font-size: 10px;")
        stage1_container.addWidget(stage1_desc)

        # Stage 1 content
        stage1_group = QGroupBox()
        stage1_layout = QVBoxLayout()

        # Processing mode selector
        mode_row = QHBoxLayout()
        mode_row.addWidget(QLabel("Processing Mode:"))
        self.processing_mode_options = {
            "JSON-First Universal (Auto-detect fields)": "json_universal",
            "ROCL fixed-width roster": "rocl",
            "ROCL fixed-width roster + Advisor column": "rocl with advisor",
            "Picture Day format (Multi-sheet)": "picture_day"
        }
        self.mode_selector = QComboBox()
        self.mode_selector.setEditable(True)
        self.mode_selector.addItems(self.processing_mode_options.keys())
        self.mode_selector.setCurrentText("JSON-First Universal (Auto-detect fields)")
        mode_row.addWidget(self.mode_selector)
        stage1_layout.addLayout(mode_row)
        
        # JSON export option
        json_row = QHBoxLayout()
        self.json_export_checkbox = QCheckBox("Export JSON for inspection")
        self.json_export_checkbox.setToolTip("Save intermediate JSON data for debugging/analysis")
        json_row.addWidget(self.json_export_checkbox)
        json_row.addStretch()
        stage1_layout.addLayout(json_row)

        # Raw source file
        raw_source_row = QHBoxLayout()
        raw_source_row.addWidget(QLabel("Raw Data File:"))
        self.raw_source_label = QLineEdit()
        self.raw_source_label.setPlaceholderText("No file selected...")
        self.raw_source_label.setReadOnly(True)
        raw_source_row.addWidget(self.raw_source_label)
        
        raw_browse_btn = QPushButton("Browse...")
        raw_browse_btn.setFixedWidth(100)
        raw_browse_btn.setStyleSheet("""
            QPushButton {
                background-color: #f0f0f0;
                color: #333;
                border: 1px solid #ddd;
                border-radius: 4px;
                padding: 5px;
            }
            QPushButton:hover {
                background-color: #e0e0e0;
                border: 1px solid #aaa;
            }
            QPushButton:pressed {
                background-color: #d0d0d0;
            }
        """)
        raw_browse_btn.clicked.connect(self.browse_raw_source)
        raw_source_row.addWidget(raw_browse_btn)
        stage1_layout.addLayout(raw_source_row)

        # Standardized output file
        std_output_row = QHBoxLayout()
        std_output_row.addWidget(QLabel("Standardized Output:"))
        self.std_output_label = QLineEdit()
        self.std_output_label.setPlaceholderText("Output file path...")
        std_output_row.addWidget(self.std_output_label)
        
        std_save_btn = QPushButton("Save As...")
        std_save_btn.setFixedWidth(100)
        std_save_btn.setStyleSheet("""
            QPushButton {
                background-color: #f0f0f0;
                color: #333;
                border: 1px solid #ddd;
                border-radius: 4px;
                padding: 5px;
            }
            QPushButton:hover {
                background-color: #e0e0e0;
                border: 1px solid #aaa;
            }
            QPushButton:pressed {
                background-color: #d0d0d0;
            }
        """)
        std_save_btn.clicked.connect(self.browse_std_output)
        std_output_row.addWidget(std_save_btn)
        stage1_layout.addLayout(std_output_row)

        # Stage 1 process button
        self.clean_btn = QPushButton("→ Clean & Standardize Data")
        self.clean_btn.setEnabled(False)
        self.clean_btn.setFixedHeight(35)
        self.clean_btn.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
                font-size: 12px;
                font-weight: bold;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #0b7dda;
            }
            QPushButton:disabled {
                background-color: #cccccc;
                color: #666666;
            }
        """)
        self.clean_btn.clicked.connect(self.process_cleaning)
        stage1_layout.addWidget(self.clean_btn)

        stage1_group.setLayout(stage1_layout)
        stage1_group.setStyleSheet("border: 1px solid #ddd; border-radius: 5px; padding: 10px;")
        stage1_container.addWidget(stage1_group)
        stage1_container.addStretch()
        
        # Add Stage 1 to horizontal layout
        two_stage_layout.addLayout(stage1_container, 1)

        # Separator
        separator = QLabel("│")
        separator.setAlignment(Qt.AlignmentFlag.AlignCenter)
        separator_font = QFont()
        separator_font.setPointSize(20)
        separator.setFont(separator_font)
        separator.setStyleSheet("color: #ddd;")
        two_stage_layout.addWidget(separator)

        # ========================================================================
        # RIGHT COLUMN: STAGE 2: CLASS SPLITTING
        # ========================================================================
        stage2_container = QVBoxLayout()

        stage2_label = QLabel("STAGE 2: Split by Class")
        stage2_font = QFont()
        stage2_font.setPointSize(11)
        stage2_font.setBold(True)
        stage2_label.setFont(stage2_font)
        stage2_label.setStyleSheet("color: #4CAF50;")
        stage2_container.addWidget(stage2_label)
       

        stage2_desc = QLabel(
            "Take standardized Excel file\nand split into separate\nsheets per class"
        )
        stage2_desc.setStyleSheet("color: #666; font-size: 10px;")
        stage2_container.addWidget(stage2_desc)

        # Stage 2 content
        stage2_group = QGroupBox()
        stage2_layout = QVBoxLayout()

        # Split source file
        split_source_row = QHBoxLayout()
        split_source_row.addWidget(QLabel("Input File:"))
        self.split_source_label = QLineEdit()
        self.split_source_label.setPlaceholderText("No file selected...")
        self.split_source_label.setReadOnly(True)
        split_source_row.addWidget(self.split_source_label)
        
        split_browse_btn = QPushButton("Browse...")
        split_browse_btn.setFixedWidth(100)
        split_browse_btn.setStyleSheet("""
            QPushButton {
                background-color: #f0f0f0;
                color: #333;
                border: 1px solid #ddd;
                border-radius: 4px;
                padding: 5px;
            }
            QPushButton:hover {
                background-color: #e0e0e0;
                border: 1px solid #aaa;
            }
            QPushButton:pressed {
                background-color: #d0d0d0;
            }
        """)
        split_browse_btn.clicked.connect(self.browse_split_source)
        split_source_row.addWidget(split_browse_btn)
        stage2_layout.addLayout(split_source_row)

        # Split output file
        split_output_row = QHBoxLayout()
        split_output_row.addWidget(QLabel("Output File:"))
        self.split_output_label = QLineEdit()
        self.split_output_label.setPlaceholderText("Output file path...")
        split_output_row.addWidget(self.split_output_label)
        
        split_save_btn = QPushButton("Save As...")
        split_save_btn.setFixedWidth(100)
        split_save_btn.setStyleSheet("""
            QPushButton {
                background-color: #f0f0f0;
                color: #333;
                border: 1px solid #ddd;
                border-radius: 4px;
                padding: 5px;
            }
            QPushButton:hover {
                background-color: #e0e0e0;
                border: 1px solid #aaa;
            }
            QPushButton:pressed {
                background-color: #d0d0d0;
            }
        """)
        split_save_btn.clicked.connect(self.browse_split_output)
        split_output_row.addWidget(split_save_btn)
        stage2_layout.addLayout(split_output_row)

        # Column mapping info
        mapping_info = QLabel("Required columns: CLASS/Class ID, Teacher, Name/Student")
        mapping_info.setStyleSheet("color: #888; font-size: 10px;")
        stage2_layout.addWidget(mapping_info)

        # Stage 2 process button
        self.split_btn = QPushButton("→ Split by Class")
        self.split_btn.setEnabled(False)
        self.split_btn.setFixedHeight(35)
        self.split_btn.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                font-size: 12px;
                font-weight: bold;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QPushButton:disabled {
                background-color: #cccccc;
                color: #666666;
            }
        """)
        self.split_btn.clicked.connect(self.process_splitting)
        stage2_layout.addWidget(self.split_btn)

        stage2_group.setLayout(stage2_layout)
        stage2_group.setStyleSheet("border: 1px solid #ddd; border-radius: 5px; padding: 10px;")
        stage2_container.addSpacing(10)  # Add 15px space above the box
        stage2_container.addWidget(stage2_group)
        stage2_container.addStretch()
        
        # Add Stage 2 to horizontal layout
        two_stage_layout.addLayout(stage2_container, 1)

        # Add two-stage layout to main layout
        main_layout.addLayout(two_stage_layout, 1)

        # ========================================================================
        # STATUS LOG
        # ========================================================================
        log_group = QGroupBox("Status Log")
        log_layout = QVBoxLayout()
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setMaximumHeight(120)
        log_layout.addWidget(self.log_text)
        log_group.setLayout(log_layout)
        main_layout.addWidget(log_group)

        main_layout.addStretch()

        self.log("Ready. Select files to begin processing.")

    # ========================================================================
    # STAGE 1: DATA CLEANING HANDLERS
    # ========================================================================

    def browse_raw_source(self):
        """Browse for raw data file from school"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select Raw Data File",
            "",
            "Excel Files (*.xlsx *.xls *.csv)"
        )
        if file_path:
            self.raw_source = file_path
            self.raw_source_label.setText(file_path)
            self.log(f"Raw data file selected: {Path(file_path).name}")
            
            # Auto-update standardized output to match the new raw file
            source_path = Path(file_path)
            suggested_output = source_path.parent / f"{source_path.stem}_STANDARDIZED.xlsx"
            self.std_output_label.setText(str(suggested_output))
            
            self.check_cleaning_ready()

    def browse_std_output(self):
        """Browse for standardized output file location"""
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Standardized File As",
            self.std_output_label.text() or "",
            "Excel Files (*.xlsx)"
        )
        if file_path:
            if not file_path.endswith('.xlsx'):
                file_path += '.xlsx'
            self.std_output_label.setText(file_path)
            self.log(f"Standardized output set: {Path(file_path).name}")
            self.check_cleaning_ready()

    def check_cleaning_ready(self):
        """Enable clean button if both files are selected"""
        ready = bool(
            self.raw_source_label.text() and 
            self.std_output_label.text()
        )
        self.clean_btn.setEnabled(ready)

    def process_cleaning(self):
        """Start Stage 1: Data cleaning process"""
        raw_file = self.raw_source_label.text()
        std_file = self.std_output_label.text()
        selected_mode = self.mode_selector.currentText().strip()
        processing_mode = self.processing_mode_options.get(selected_mode, selected_mode)

        if not raw_file or not std_file:
            QMessageBox.warning(self, "Missing Files", "Please select both raw data and output files.")
            return

        # Check for overwrite
        if Path(std_file).exists():
            reply = QMessageBox.question(
                self,
                "File Exists",
                "Output file already exists. Overwrite it?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if reply == QMessageBox.StandardButton.No:
                return

        self.clean_btn.setEnabled(False)
        self.log_text.clear()
        self.log(f"[STAGE 1] Starting {selected_mode} processing...")

        # Get pipeline based on processing mode
        pipeline = self._get_processing_pipeline(processing_mode, raw_file, std_file)
        
        self.clean_thread = ProcessThread(raw_file, std_file, pipeline)
        self.clean_thread.progress.connect(self.log)
        self.clean_thread.finished.connect(self.on_cleaning_finished)
        self.clean_thread.start()
    
    def _get_processing_pipeline(self, processing_mode: str, raw_file: str, std_file: str):
        """Get appropriate pipeline based on processing mode"""
        if processing_mode == "json_universal":
            # JSON-First Universal mode - use new dedicated cleaning stage
            json_export = self.json_export_checkbox.isChecked()
            json_path = None
            if json_export:
                json_path = str(Path(std_file).with_suffix('.json'))
            
            from cleaning_stages import get_cleaning_only_pipeline
            pipeline = get_cleaning_only_pipeline("json_universal")
            # Set JSON export path on the stage
            if pipeline.stages:
                pipeline.stages[0].json_output_path = json_path
            return pipeline
        
        elif processing_mode == "picture_day":
            # Picture Day format
            try:
                from picture_day_cleaning_stage import PictureDayCleaningStage
                return ProcessingPipeline(stages=[PictureDayCleaningStage()])
            except ImportError:
                self.log("Picture Day cleaning stage not available, using universal mode")
                from cleaning_stages import get_cleaning_only_pipeline
                return get_cleaning_only_pipeline("json_universal")
        
        else:
            # Legacy ROCL formats
            try:
                from cleaning_stages import get_cleaning_only_pipeline
                return get_cleaning_only_pipeline(processing_mode)
            except ImportError:
                self.log("ROCL cleaning stages not available, using universal mode")
                from cleaning_stages import get_cleaning_only_pipeline
                return get_cleaning_only_pipeline("json_universal")

    def on_cleaning_finished(self, success, message):
        """Handle Stage 1 completion"""
        self.log(message)
        
        if success:
            QMessageBox.information(self, "Cleaning Complete", message)
            
            # Auto-populate Stage 2 input with Stage 1 output
            std_file = self.std_output_label.text()
            self.split_source_label.setText(std_file)
            
            # Auto-suggest Stage 2 output
            source_path = Path(std_file)
            suggested_output = source_path.parent / f"{source_path.stem}_BY_CLASS.xlsx"
            self.split_output_label.setText(str(suggested_output))
            
            self.check_splitting_ready()
            
            reply = QMessageBox.question(
                self,
                "Open File?",
                "Would you like to open the standardized file?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if reply == QMessageBox.StandardButton.Yes:
                self.open_file(std_file)
        else:
            QMessageBox.critical(self, "Cleaning Failed", message)
        
        self.clean_btn.setEnabled(True)

    # ========================================================================
    # STAGE 2: CLASS SPLITTING HANDLERS
    # ========================================================================

    def browse_split_source(self):
        """Browse for standardized input file"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select Standardized Data File",
            "",
            "Excel Files (*.xlsx *.xls)"
        )
        if file_path:
            self.split_source_label.setText(file_path)
            self.log(f"Split input file selected: {Path(file_path).name}")
            
            # Also update Stage 1 standardized output if empty
            if not self.std_output_label.text():
                self.std_output_label.setText(file_path)
                self.log(f"Stage 1 standardized output updated: {Path(file_path).name}")
            
            # Auto-update output for Stage 2 to match the new input file
            source_path = Path(file_path)
            suggested_output = source_path.parent / f"{source_path.stem}_BY_CLASS.xlsx"
            self.split_output_label.setText(str(suggested_output))
            
            self.check_splitting_ready()

    def browse_split_output(self):
        """Browse for split output file location"""
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Split File As",
            self.split_output_label.text() or "",
            "Excel Files (*.xlsx)"
        )
        if file_path:
            if not file_path.endswith('.xlsx'):
                file_path += '.xlsx'
            self.split_output_label.setText(file_path)
            self.log(f"Split output set: {Path(file_path).name}")
            self.check_splitting_ready()

    def check_splitting_ready(self):
        """Enable split button if both files are selected"""
        ready = bool(
            self.split_source_label.text() and 
            self.split_output_label.text()
        )
        self.split_btn.setEnabled(ready)

    def process_splitting(self):
        """Start Stage 2: Class splitting process"""
        split_file = self.split_source_label.text()
        output_file = self.split_output_label.text()

        if not split_file or not output_file:
            QMessageBox.warning(self, "Missing Files", "Please select both input and output files.")
            return

        # Check for overwrite
        if Path(output_file).exists():
            reply = QMessageBox.question(
                self,
                "File Exists",
                "Output file already exists. Overwrite it?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if reply == QMessageBox.StandardButton.No:
                return

        self.split_btn.setEnabled(False)
        self.log_text.clear()
        self.log("[STAGE 2] Starting class splitting...")

        # Use only ClassSplittingStage for this process
        splitting_pipeline = ProcessingPipeline(stages=[ClassSplittingStage()])
        
        self.split_thread = ProcessThread(split_file, output_file, splitting_pipeline)
        self.split_thread.progress.connect(self.log)
        self.split_thread.finished.connect(self.on_splitting_finished)
        self.split_thread.start()

    def on_splitting_finished(self, success, message):
        """Handle Stage 2 completion"""
        self.log(message)
        
        if success:
            QMessageBox.information(self, "Splitting Complete", message)
            
            reply = QMessageBox.question(
                self,
                "Open File?",
                "Would you like to open the split file?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if reply == QMessageBox.StandardButton.Yes:
                self.open_file(self.split_output_label.text())
        else:
            QMessageBox.critical(self, "Splitting Failed", message)
        
        self.split_btn.setEnabled(True)

    # ========================================================================
    # UTILITY METHODS
    # ========================================================================

    def open_file(self, file_path):
        """Open file with default application"""
        import subprocess
        import platform
        
        try:
            if platform.system() == 'Darwin':  # macOS
                subprocess.run(['open', file_path])
            elif platform.system() == 'Windows':
                subprocess.run(['start', '', file_path], shell=True)
            else:  # Linux
                subprocess.run(['xdg-open', file_path])
        except Exception as e:
            self.log(f"Could not open file: {str(e)}")

    def log(self, message):
        self.log_text.append(message)


def main():
    app = QApplication(sys.argv)
    
    # Set application style
    app.setStyle('Fusion')
    
    window = ClassRosterGUI()
    window.show()
    
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
