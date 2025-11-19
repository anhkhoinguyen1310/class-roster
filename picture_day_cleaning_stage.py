"""
Picture Day Format Cleaning Stage

This format has:
- Multiple sheets (one per grade: 6th, 7th, 8th)
- Student data with contact info (multiple contacts per student)
- Separate schedule sheet with class-teacher mappings
- No direct class assignment for students

Strategy:
1. Extract all data to normalized JSON structure
2. Map students to classes based on grade and schedule
3. Handle missing teacher info gracefully
4. Output: Teacher, Student Name, Class (or Grade if no class)
"""

import json
from openpyxl import Workbook
from collections import defaultdict

# Import base class
try:
    from class_roster_ui import DataCleaningStage
    BASE_CLASS_AVAILABLE = True
except ImportError:
    # Fallback if not running in the main app
    BASE_CLASS_AVAILABLE = False
    class DataCleaningStage:
        def process(self, data: dict) -> dict:
            pass
        def get_stage_name(self) -> str:
            pass


class PictureDayCleaningStage(DataCleaningStage):
    """
    Custom cleaning for Picture Day format.
    
    Features:
    - Multi-sheet processing (one sheet per grade)
    - JSON-based extraction and normalization
    - Smart class-teacher mapping from schedule
    - Handles students with multiple contacts
    - Works without teacher info (optional)
    """
    
    def __init__(self, include_contacts=False):
        """
        Args:
            include_contacts: If True, include contact info in output
        """
        self.include_contacts = include_contacts
        self.json_data = None
    
    def process(self, data: dict) -> dict:
        """Clean and reformat Picture Day data"""
        wb_in = data.get("workbook")
        
        if not wb_in:
            raise ValueError("Input workbook not found")
        
        # Step 1: Extract all data to JSON
        self.json_data = self._extract_to_json(wb_in)
        
        # Step 2: Map classes and teachers
        self._map_classes_and_teachers(wb_in)
        
        # Step 3: Create cleaned workbook
        wb_out = self._create_cleaned_workbook()
        
        # Update data dict for next stage
        data["workbook"] = wb_out
        data["worksheet"] = wb_out.active
        data["cleaned_record_count"] = len(self.json_data['students'])
        data["json_extraction"] = self.json_data  # Keep JSON for reference
        
        return data
    
    def _extract_to_json(self, workbook):
        """
        Extract all student data to JSON structure.
        
        Returns:
            {
                'students': [
                    {
                        'first_name': str,
                        'last_name': str,
                        'full_name': str,
                        'grade': str,
                        'class': str (initially None),
                        'teacher': str (initially None),
                        'contacts': [...]
                    }
                ],
                'schedule': {...}
            }
        """
        json_structure = {
            'students': [],
            'schedule': {},
            'metadata': {
                'total_sheets': 0,
                'grades_processed': []
            }
        }
        
        # Define which sheets contain student data (exclude Schedule sheet)
        processed_students = {}  # Track unique students by (first, last, grade)
        
        for sheet_name in workbook.sheetnames:
            # Skip schedule sheet
            if 'schedule' in sheet_name.lower():
                continue
                
            # Check if this is a grade sheet
            if any(grade in sheet_name for grade in ['6th', '7th', '8th', 'Grade']):
                ws = workbook[sheet_name]
                json_structure['metadata']['total_sheets'] += 1
                json_structure['metadata']['grades_processed'].append(sheet_name)
                
                # Extract students from this sheet
                # Column structure: First Name (0), Last Name (1), Grade (2), Contact info starts at (3)
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or not row[0]:  # Skip empty rows
                        continue
                    
                    student_first = self._clean_text(row[0]) if row[0] else ""
                    student_last = self._clean_text(row[1]) if row[1] else ""
                    grade = self._clean_text(row[2]) if row[2] else ""
                    
                    if not student_first or not student_last:
                        continue
                    
                    # Create unique key for student (deduplicate if same student appears multiple times)
                    student_key = f"{student_first.upper()}|{student_last.upper()}|{grade}"
                    
                    # Create or update student record
                    if student_key not in processed_students:
                        # IMPORTANT: Combine first + last name properly
                        full_name = f"{student_first} {student_last}".strip()
                        processed_students[student_key] = {
                            'first_name': student_first,
                            'last_name': student_last,
                            'full_name': full_name,
                            'grade': grade,
                            'class': None,  # Will be assigned later
                            'teacher': None,  # Will be assigned later
                            'contacts': []
                        }
                    
                    # Add contact info if present (columns 3-7)
                    if len(row) >= 8:
                        contact_first = self._clean_text(row[3]) if row[3] else ""
                        contact_last = self._clean_text(row[4]) if row[4] else ""
                        relationship = self._clean_text(row[5]) if row[5] else ""
                        phone = self._clean_phone(row[6]) if row[6] else ""
                        email = self._clean_text(row[7]) if len(row) > 7 and row[7] else ""
                        
                        if contact_first or contact_last or email:
                            processed_students[student_key]['contacts'].append({
                                'first_name': contact_first,
                                'last_name': contact_last,
                                'relationship': relationship,
                                'phone': phone,
                                'email': email
                            })
        
        # Convert to list
        json_structure['students'] = list(processed_students.values())
        
        return json_structure
    
    def _map_classes_and_teachers(self, workbook):
        """
        Extract schedule data and map to students.
        Maps grade levels to classes and teachers.
        Handles teacher names with special formats (newlines, slashes, etc.)
        """
        # Find schedule sheet (usually named "Schedule " with trailing space)
        schedule_sheet = None
        for sheet_name in workbook.sheetnames:
            if 'schedule' in sheet_name.lower():
                schedule_sheet = workbook[sheet_name]
                break
        
        if not schedule_sheet:
            # No schedule sheet, just use grade as class
            for student in self.json_data['students']:
                student['class'] = student['grade']
                student['teacher'] = ""  # No teacher info available
            return
        
        # Extract class-teacher mappings from schedule
        # Header is in row 4: Column B=Class, Column E=Teacher
        class_teacher_map = {}
        
        for row in schedule_sheet.iter_rows(min_row=5, values_only=True):  # Start from row 5 (after header)
            if not row or len(row) < 5:
                continue
            
            class_val = row[1]  # Column B (index 1) = Class
            teacher = row[4]     # Column E (index 4) = Teacher
            
            if class_val and teacher:
                class_str = str(class_val).strip()
                teacher_str = str(teacher).strip()
                
                # Clean teacher name: handle newlines, slashes, trim whitespace
                teacher_str = self._clean_teacher_name(teacher_str)
                
                if not teacher_str:  # Skip if teacher name is empty after cleaning
                    continue
                
                # Parse class value (e.g., "8.1", "7.3", "6.2" or "8.3 / 8.1")
                if '/' in class_str:
                    # Handle "8.3 / 8.1" format - split by slash
                    classes = [c.strip() for c in class_str.split('/')]
                    for c in classes:
                        if c:  # Skip empty strings
                            class_teacher_map[c] = teacher_str
                else:
                    class_teacher_map[class_str] = teacher_str
        
        self.json_data['schedule'] = class_teacher_map
        
        # Map students to classes (distribute evenly across grade classes)
        from collections import defaultdict
        grade_classes = defaultdict(list)
        for class_name in class_teacher_map.keys():
            # Extract grade from class name (e.g., "8.1" -> "8")
            grade_prefix = class_name.split('.')[0] if '.' in str(class_name) else str(class_name)[0]
            if grade_prefix == '6':
                grade_classes['6th Grade'].append(class_name)
            elif grade_prefix == '7':
                grade_classes['7th Grade'].append(class_name)
            elif grade_prefix == '8':
                grade_classes['8th Grade'].append(class_name)
        
        # Sort students by grade
        students_by_grade = defaultdict(list)
        for student in self.json_data['students']:
            grade_key = student['grade']
            # Normalize grade key
            if '6th' in grade_key or '6 ' in grade_key:
                students_by_grade['6th Grade'].append(student)
            elif '7th' in grade_key or '7 ' in grade_key:
                students_by_grade['7th Grade'].append(student)
            elif '8th' in grade_key or '8 ' in grade_key:
                students_by_grade['8th Grade'].append(student)
        
        # Assign classes to students (round-robin distribution)
        for grade, students in students_by_grade.items():
            classes = grade_classes.get(grade, [])
            
            if not classes:
                # No classes found for this grade, use grade as class
                for student in students:
                    student['class'] = grade
                    student['teacher'] = ""
            else:
                # Distribute students across available classes
                for i, student in enumerate(students):
                    class_name = classes[i % len(classes)]
                    student['class'] = class_name
                    student['teacher'] = class_teacher_map.get(class_name, "")
    
    def _clean_text(self, text):
        """Clean text: trim and title case"""
        if not text:
            return ""
        
        cleaned = str(text).strip()
        
        # Convert to title case if all caps
        if cleaned.isupper():
            cleaned = cleaned.title()
        
        return cleaned
    
    def _clean_teacher_name(self, teacher_str):
        """
        Clean teacher names that may have special formats:
        - "Reno/ Orio " -> "Reno/Orio"
        - "Ms.Roy\nMs.Francois\nMs.Muhammad" -> "Ms. Roy, Ms. Francois, Ms. Muhammad"
        - "Ms. Roy" -> "Ms. Roy"
        
        Returns cleaned teacher name(s)
        """
        if not teacher_str:
            return ""
        
        # Replace newlines with commas and space
        teacher_str = teacher_str.replace('\n', ', ')
        
        # Clean up slashes - keep them but trim whitespace around them
        # "Reno/ Orio " -> "Reno / Orio"
        parts = teacher_str.split('/')
        parts = [p.strip() for p in parts]
        teacher_str = ' / '.join(parts)
        
        # Split by comma to handle multiple teachers
        teachers = [t.strip() for t in teacher_str.split(',')]
        
        # Clean each teacher name
        cleaned_teachers = []
        for t in teachers:
            if t:  # Only add non-empty names
                # Title case the name if needed
                if t.isupper():
                    t = t.title()
                cleaned_teachers.append(t)
        
        # Join back with commas
        result = ', '.join(cleaned_teachers)
        
        return result.strip()
    
    def _clean_phone(self, phone):
        """Clean phone number"""
        if not phone:
            return ""
        
        # Convert to string and clean
        phone_str = str(phone).strip()
        
        # Remove .0 from floats
        if phone_str.endswith('.0'):
            phone_str = phone_str[:-2]
        
        return phone_str
    def _create_cleaned_workbook(self):
        """
        Create new workbook with cleaned data.

        Output is intentionally minimal and standardized:
        - Teacher
        - Student Name
        - Class  (falls back to grade if no class was assigned)
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Cleaned Data"

        # Unified header: same shape as ROCL / others
        ws.append(["Teacher", "Student Name", "Class"])

        # Write student records – still sorted by (grade, last name) for stability
        for student in sorted(self.json_data['students'], key=lambda s: (s['grade'], s['last_name'])):
            teacher = student.get('teacher', '')
            student_name = student['full_name']
            # If no class was mapped, fall back to grade
            class_val = student.get('class', student['grade'])

            ws.append([teacher, student_name, class_val])

        return wb

    
    def get_stage_name(self) -> str:
        return "Data Cleaning (Picture Day Format)"
    
    def export_json(self, filepath):
        """Export the extracted JSON for inspection"""
        if self.json_data:
            with open(filepath, 'w') as f:
                json.dump(self.json_data, f, indent=2)


# ============================================================================
# STANDALONE TEST
# ============================================================================

if __name__ == "__main__":
    print("Picture Day Cleaning Stage - Test Mode")
    print("=" * 70)
    
    from openpyxl import load_workbook
    import os
    
    test_file = "/mnt/user-data/uploads/_Picture_Day_Student_Information_and_schedule_25_.xlsx"
    
    if os.path.exists(test_file):
        print(f"Loading test file: {test_file}\n")
        wb = load_workbook(test_file, data_only=True)
        
        data = {"workbook": wb}
        
        # Run cleaning stage
        stage = PictureDayCleaningStage(include_contacts=False)
        print(f"Running: {stage.get_stage_name()}\n")
        
        cleaned_data = stage.process(data)
        
        print(f"Cleaning complete!")
        print(f"Records processed: {cleaned_data.get('cleaned_record_count', 0)}")
        
        # Export JSON for inspection
        json_path = "/home/claude/picture_day_extracted.json"
        stage.export_json(json_path)
        print(f"JSON exported to: {json_path}")
        
        # Show sample of cleaned data
        ws_out = cleaned_data["worksheet"]
        print("\nFirst 15 cleaned records:")
        print(f"{'Teacher':<25} | {'Student Name':<25} | {'Class':<10} | {'Grade':<10}")
        print("-" * 80)
        for i, row in enumerate(ws_out.iter_rows(min_row=2, max_row=16, values_only=True), 1):
            print(f"{str(row[0]):<25} | {str(row[1]):<25} | {str(row[2]):<10} | {str(row[3]):<10}")
        
        # Show class distribution
        print("\n" + "=" * 70)
        print("Class Distribution:")
        class_counts = defaultdict(int)
        for row in ws_out.iter_rows(min_row=2, values_only=True):
            if row[2]:
                class_counts[row[2]] += 1
        
        for class_name, count in sorted(class_counts.items()):
            teacher = ""
            for row in ws_out.iter_rows(min_row=2, values_only=True):
                if row[2] == class_name and row[0]:
                    teacher = row[0]
                    break
            print(f"  {class_name}: {count} students (Teacher: {teacher})")
        
        # Save test output
        output_path = "/home/claude/test_picture_day_cleaned.xlsx"
        cleaned_data["workbook"].save(output_path)
        print(f"\n✓ Test output saved to: {output_path}")
    else:
        print(f"Test file not found: {test_file}")
