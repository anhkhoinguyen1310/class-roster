#!/usr/bin/env python3
"""
JSON-First Excel Processor Demo

This demonstrates the new JSON-first processing workflow:
1. Scan all XLSX documents (including multiple sheets)
2. Extract everything as JSON objects
3. Match to correct fields
4. Return to correct XLSX file

Usage:
    python json_processor_demo.py input_file.xlsx output_file.xlsx
"""

import sys
import json
from pathlib import Path
from openpyxl import load_workbook
from class_roster_ui import DataCleaningStage, ProcessingPipeline


def demo_json_first_processing(input_file: str, output_file: str, export_json: bool = True):
    """
    Demonstrate the JSON-first processing workflow
    
    Args:
        input_file: Path to input Excel file
        output_file: Path to save processed Excel file
        export_json: Whether to export intermediate JSON
    """
    print("=" * 80)
    print("JSON-First Excel Processor Demo")
    print("=" * 80)
    
    # Setup paths
    input_path = Path(input_file)
    output_path = Path(output_file)
    json_path = output_path.with_suffix('.json') if export_json else None
    
    if not input_path.exists():
        print(f"❌ Error: Input file not found: {input_path}")
        return False
    
    print(f"📁 Input file:  {input_path}")
    print(f"📁 Output file: {output_path}")
    if json_path:
        print(f"📁 JSON export: {json_path}")
    print()
    
    try:
        # Load workbook
        print("📖 Loading workbook...")
        wb = load_workbook(input_path, data_only=True)
        print(f"   Sheets found: {wb.sheetnames}")
        print()
        
        # Create JSON-first processing pipeline
        print("🔄 Setting up JSON-first processing pipeline...")
        pipeline = ProcessingPipeline(stages=[
            DataCleaningStage(
                export_json=export_json,
                json_output_path=str(json_path) if json_path else None
            )
        ])
        
        # Process the file
        print("⚡ Processing file...")
        success, message, result_data = pipeline.execute(
            str(input_path),
            str(output_path),
            progress_callback=lambda msg: print(f"   {msg}")
        )
        
        if success:
            print(f"✅ {message}")
            
            # Show processing results
            if 'normalized_records' in result_data:
                records = result_data['normalized_records']
                print(f"📊 Processed {len(records)} records")
                
                # Show sample records
                print("\n📋 Sample processed records:")
                print("-" * 80)
                print(f"{'Student Name':<25} | {'Class':<15} | {'Teacher':<20} | {'Source':<15}")
                print("-" * 80)
                
                for i, record in enumerate(records[:10]):  # Show first 10
                    print(f"{record.get('student_name', ''):<25} | "
                          f"{record.get('class_id', ''):<15} | "
                          f"{record.get('teacher', ''):<20} | "
                          f"{record.get('source_sheet', ''):<15}")
                
                if len(records) > 10:
                    print(f"... and {len(records) - 10} more records")
            
            # Show JSON info if exported
            if json_path and json_path.exists():
                print(f"\n📄 JSON exported successfully to: {json_path}")
                
                # Show JSON summary
                with open(json_path, 'r') as f:
                    json_data = json.load(f)
                
                print("   JSON Summary:")
                print(f"   - Total sheets processed: {json_data.get('extraction_summary', {}).get('total_sheets', 0)}")
                print(f"   - Data sheets: {len(json_data.get('raw_data', {}))}")
                print(f"   - Normalized records: {json_data.get('record_count', 0)}")
        
        else:
            print(f"❌ Processing failed: {message}")
            return False
        
        print()
        print("=" * 80)
        print("✅ Demo completed successfully!")
        return True
        
    except Exception as e:
        print(f"❌ Error during processing: {str(e)}")
        return False


def main():
    """Main entry point"""
    if len(sys.argv) < 3:
        print("Usage: python json_processor_demo.py input_file.xlsx output_file.xlsx")
        print()
        print("This demo shows the JSON-first processing workflow:")
        print("1. Scans all sheets in the Excel file")
        print("2. Extracts data to JSON format")
        print("3. Matches fields intelligently")
        print("4. Outputs standardized Excel file")
        print("5. Optionally exports JSON for inspection")
        return
    
    input_file = sys.argv[1]
    output_file = sys.argv[2]
    
    success = demo_json_first_processing(input_file, output_file)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()